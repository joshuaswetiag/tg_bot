import csv
import re
from io import BytesIO, StringIO

_HEADER_VALUES = frozenset({"email", "login", "username", "user", "password"})


def normalize_account_pair(login: str, password: str) -> str | None:
    login = login.strip()
    password = password.strip()
    if not login or not password:
        return None
    if login.lower() in _HEADER_VALUES or password.lower() in _HEADER_VALUES:
        return None
    return f"{login}:{password}"


def normalize_account_line(line: str) -> str | None:
    line = line.strip()
    if not line or line.startswith("#"):
        return None
    if ":" not in line:
        return None
    login, password = line.split(":", 1)
    return normalize_account_pair(login, password)


def _collect_accounts(rows: list[tuple[str, str]]) -> list[str]:
    seen: set[str] = set()
    accounts: list[str] = []
    for login, password in rows:
        normalized = normalize_account_pair(login, password)
        if normalized and normalized not in seen:
            seen.add(normalized)
            accounts.append(normalized)
    return accounts


def _rows_from_sheet_rows(sheet_rows) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for row in sheet_rows:
        if not row or len(row) < 3:
            continue
        login_raw = row[1]
        password_raw = row[2]
        if login_raw is None or password_raw is None:
            continue
        pairs.append((str(login_raw), str(password_raw)))
    return pairs


def parse_accounts_from_text(text: str) -> list[str]:
    accounts: list[str] = []
    seen: set[str] = set()
    for line in text.splitlines():
        normalized = normalize_account_line(line)
        if normalized and normalized not in seen:
            seen.add(normalized)
            accounts.append(normalized)
    return accounts


def parse_accounts_from_csv(data: bytes) -> list[str]:
    text = data.decode("utf-8-sig", errors="ignore")
    if not text.strip():
        return []

    try:
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(StringIO(text), dialect)
    pairs: list[tuple[str, str]] = []
    for row in reader:
        if len(row) < 3:
            continue
        pairs.append((row[1], row[2]))
    return _collect_accounts(pairs)


def parse_accounts_from_excel(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        return _collect_accounts(_rows_from_sheet_rows(ws.iter_rows(min_row=1, values_only=True)))
    finally:
        wb.close()


def parse_accounts_upload(filename: str, data: bytes) -> list[str]:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return parse_accounts_from_excel(data)
    if lower.endswith(".csv"):
        return parse_accounts_from_csv(data)
    return parse_accounts_from_text(data.decode("utf-8", errors="ignore"))


def format_accounts_delivery_file(account_lines: list[str]) -> str:
    blocks: list[str] = []
    for line in account_lines:
        normalized = normalize_account_line(line)
        if not normalized:
            continue
        login, password = normalized.split(":", 1)
        label = "Email" if "@" in login else "Login"
        blocks.append(f"{label}: {login}\nPassword: {password}")
    return "\n\n".join(blocks) + ("\n" if blocks else "")


def account_count_label(count: int) -> str:
    return f"{count} proxy account{'s' if count != 1 else ''}"
